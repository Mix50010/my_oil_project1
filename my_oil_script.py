import platform, psutil
from screeninfo import get_monitors
from openpyxl import Workbook

data = dict()

data['os_version'] = platform.platform()
data['pc_name'] = platform.node()
data['cpu'] = platform.machine()
data['ram_size'] = round(psutil.virtual_memory().total / 1024 ** 3, 1)
for i, monitor in enumerate(get_monitors()):
    data[f'monitor_{i}'] = str((monitor.width, monitor.height))

print(*map(lambda x: f'{x}: {data[x]}', data), sep='\n')

wb = Workbook()
ws = wb.active
ws.title = "Platform info"
for i, pair in enumerate(data.items(), 1):
    key, value = pair
    ws[f'A{i}'] = key
    ws[f'B{i}'] = value
wb.save('platform_info.xlsx')
print("platform_info.xlsx was created and now it is located in exe directory")
input('Press ENTER to escape')